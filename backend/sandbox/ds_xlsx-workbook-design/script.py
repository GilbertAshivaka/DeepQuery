import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference

# ── Palette & helpers ────────────────────────────────────────────────────────
DARK = "1B5E20"
MID = "2E7D32"
LIGHT = "E8F5E9"
WHITE = "FFFFFF"
GREY1 = "F5F5F5"
INK = "212121"
LINE = "BBBBBB"

def _font(bold=False, color=INK, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def _align(where="left"):
    return Alignment(horizontal=where, vertical="center", wrap_text=(where != "right"))

def _border():
    s = Side(style="thin", color=LINE)
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, value, *, bold=False, bg=None, fg=INK, align="left", size=10, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold, fg, size)
    c.alignment = _align(align)
    c.border = _border()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt
    return c

# ── Data definition ────────────────────────────────────────────────────────
categories = [
    "Rent",
    "Utilities",
    "Supplies",
    "Labor",
    "Marketing",
    "Equipment Maintenance",
    "Insurance",
    "Miscellaneous"
]

# amounts per month (Jan, Feb, Mar) – arbitrary realistic numbers
jan_vals = [1200, 300, 450, 2500, 200, 150, 180, 100]
feb_vals = [1200, 280, 470, 2600, 250, 130, 180, 120]
mar_vals = [1200, 310, 430, 2550, 220, 160, 180, 110]

# compute totals
monthly_totals = [
    sum(jan_vals),
    sum(feb_vals),
    sum(mar_vals)
]
category_totals = [jan + feb + mar for jan, feb, mar in zip(jan_vals, feb_vals, mar_vals)]
grand_total = sum(monthly_totals)

# ── Workbook construction ───────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Summary"
ws.sheet_view.showGridLines = False

# Title block
ws.merge_cells("B2:F2")
title_cell = ws["B2"]
title_cell.value = "Small Cafe Budget"
title_cell.font = _font(bold=True, color=WHITE, size=24)
title_cell.alignment = _align("center")
title_cell.fill = PatternFill("solid", fgColor=DARK)

ws.merge_cells("B3:F3")
subtitle_cell = ws["B3"]
subtitle_cell.value = "Budget Overview – Jan to Mar 2023"
subtitle_cell.font = _font(italic=True, color=INK, size=12)
subtitle_cell.alignment = _align("center")

# Header row (row 4)
headers = ["Category", "Jan", "Feb", "Mar", "Total"]
for col_idx, header in enumerate(headers, start=1):
    cell(ws, 4, col_idx, header, bold=True, bg=DARK, fg=WHITE, align="center", size=12)

# Data rows (starting at row 5)
start_row = 5
for i, cat in enumerate(categories):
    row = start_row + i
    bg_fill = GREY1 if i % 2 else WHITE
    # Category name (bold)
    cell(ws, row, 1, cat, bold=True, bg=bg_fill, fg=INK, align="left")
    # Jan, Feb, Mar values
    cell(ws, row, 2, jan_vals[i], bg=bg_fill, fg=INK, align="right", num_fmt="#,##0")
    cell(ws, row, 3, feb_vals[i], bg=bg_fill, fg=INK, align="right", num_fmt="#,##0")
    cell(ws, row, 4, mar_vals[i], bg=bg_fill, fg=INK, align="right", num_fmt="#,##0")
    # Category total
    cell(ws, row, 5, category_totals[i], bg=bg_fill, fg=INK, align="right", num_fmt="#,##0")

# Totals row
totals_row = start_row + len(categories)
cell(ws, totals_row, 1, "Total", bold=True, bg=DARK, fg=WHITE, align="right", size=12)
cell(ws, totals_row, 2, monthly_totals[0], bold=True, bg=DARK, fg=WHITE, align="right", size=12, num_fmt="#,##0")
cell(ws, totals_row, 3, monthly_totals[1], bold=True, bg=DARK, fg=WHITE, align="right", size=12, num_fmt="#,##0")
cell(ws, totals_row, 4, monthly_totals[2], bold=True, bg=DARK, fg=WHITE, align="right", size=12, num_fmt="#,##0")
cell(ws, totals_row, 5, grand_total, bold=True, bg=DARK, fg=WHITE, align="right", size=12, num_fmt="#,##0")

# Freeze panes (keep title+header visible)
ws.freeze_panes = "A5"

# Column widths
col_widths = {
    1: 25,  # Category
    2: 12,  # Jan
    3: 12,  # Feb
    4: 12,  # Mar
    5: 14   # Total
}
for col_idx, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Conditional formatting on Jan column (highlight low‑high values)
cs_range = f"B{start_row}:B{totals_row-1}"
ws.conditional_formatting.add(cs_range,
    ColorScaleRule(start_type="min", start_color=LIGHT,
                   end_type="max", end_color=DARK))

# Chart – Monthly totals bar chart
chart = BarChart()
chart.type = "col"
chart.title = "Monthly Expenditure"
chart.style = 10
chart.y_axis.title = "Amount (USD)"
chart.x_axis.title = "Month"

data = Reference(ws, min_col=2, max_col=4, min_row=totals_row, max_row=totals_row)
cats = Reference(ws, min_col=2, max_col=4, min_row=4, max_row=4)  # header names as categories
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)
chart.height = 7
chart.width = 12
ws.add_chart(chart, "G5")

# Save workbook
output_path = "/workspace/output/cafe_budget.xlsx"
wb.save(output_path)
print("done")