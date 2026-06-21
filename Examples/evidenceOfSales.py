import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import datetime

wb = openpyxl.Workbook()

# ── COLOURS ──────────────────────────────────────────────────────────────────
G_DARK   = "1B5E20"
G_MID    = "2E7D32"
G_LIGHT  = "E8F5E9"
G_MLIGHT = "C8E6C9"
GOLD     = "F9A825"
GOLD_BG  = "FFF8E1"
WHITE    = "FFFFFF"
GREY1    = "F5F5F5"
GREY2    = "EEEEEE"
DARK     = "212121"
RED_SOFT = "FFEBEE"
BLUE_S   = "E3F2FD"
PURPLE_S = "F3E5F5"

def hdr_fill(hex_):  return PatternFill("solid", fgColor=hex_)
def font(bold=False, color="212121", size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def right():  return Alignment(horizontal="right",  vertical="center")
def thin_border(sides="all"):
    s = Side(style="thin", color="BBBBBB")
    n = Side(style=None)
    if sides == "all":  return Border(left=s, right=s, top=s, bottom=s)
    if sides == "top":  return Border(top=s)
    if sides == "bot":  return Border(bottom=s)
    return Border(left=s, right=s, top=s, bottom=s)
def thick_bottom():
    return Border(bottom=Side(style="medium", color=G_MID))

def style_cell(ws, row, col, value, bold=False, bg=None, fg=DARK,
               align="left", size=10, italic=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold, fg, size, italic)
    c.alignment = center() if align=="center" else (right() if align=="right" else left())
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.border = thin_border()
    if num_fmt: c.number_format = num_fmt
    return c

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER / SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "Summary"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 2
for col in ["B","C","D","E","F","G","H"]:
    ws0.column_dimensions[col].width = 20

# Title block
ws0.merge_cells("B2:H2")
c = ws0["B2"]
c.value = "REFORMED CREATIONS"
c.font = Font(name="Georgia", bold=True, size=28, color=WHITE)
c.fill = PatternFill("solid", fgColor=G_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws0.row_dimensions[2].height = 45

ws0.merge_cells("B3:H3")
c = ws0["B3"]
c.value = "Sales & Transaction Evidence Sheet  |  November 2024 – May 2026"
c.font = Font(name="Calibri", bold=False, size=13, color=G_DARK, italic=True)
c.fill = PatternFill("solid", fgColor=G_LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center")
ws0.row_dimensions[3].height = 28

ws0.merge_cells("B4:H4")
c = ws0["B4"]
c.value = "Kilifi, Kenya  |  samaa@reformedcreations.co.ke  |  Prepared: May 2026"
c.font = Font(name="Calibri", size=10, color="666666")
c.fill = PatternFill("solid", fgColor=G_LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center")
ws0.row_dimensions[4].height = 20

# Business details
details = [
    ("Business Name:", "Reformed Creations"),
    ("Founders:", "Samaa Shariff & Fridah Santa"),
    ("Location:", "Kilifi, Kilifi County, Kenya"),
    ("Trading Since:", "November 2024"),
    ("Business Type:", "Social Enterprise — Eco-Fashion"),
    ("Document Purpose:", "Application Evidence of Sales & Operations"),
]
ws0.row_dimensions[5].height = 8
for i, (lbl, val) in enumerate(details):
    r = 6 + i
    ws0.merge_cells(f"B{r}:C{r}")
    ws0[f"B{r}"].value = lbl
    ws0[f"B{r}"].font = Font(name="Calibri", bold=True, size=10, color=G_DARK)
    ws0[f"B{r}"].fill = PatternFill("solid", fgColor=G_MLIGHT)
    ws0[f"B{r}"].alignment = left()
    ws0[f"B{r}"].border = thin_border()
    ws0.merge_cells(f"D{r}:H{r}")
    ws0[f"D{r}"].value = val
    ws0[f"D{r}"].font = Font(name="Calibri", size=10, color=DARK)
    ws0[f"D{r}"].fill = PatternFill("solid", fgColor=WHITE)
    ws0[f"D{r}"].alignment = left()
    ws0[f"D{r}"].border = thin_border()
    ws0.row_dimensions[r].height = 20

# Monthly summary table
ws0.row_dimensions[13].height = 12
ws0.merge_cells("B14:H14")
c = ws0["B14"]
c.value = "MONTHLY REVENUE SUMMARY"
c.font = Font(name="Georgia", bold=True, size=12, color=WHITE)
c.fill = PatternFill("solid", fgColor=G_MID)
c.alignment = Alignment(horizontal="center", vertical="center")
ws0.row_dimensions[14].height = 22

monthly_hdrs = ["Period","Revenue (KES)","Units Sold","Jewellery (KES)","Key Holders (KES)","Charms (KES)","Notes"]
monthly_cols = ["B","C","D","E","F","G","H"]
for ci, h in enumerate(monthly_hdrs):
    c = ws0.cell(row=15, column=ord(monthly_cols[ci])-64)
    c.value = h
    c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=G_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
ws0.row_dimensions[15].height = 28

monthly_data = [
    ("Nov 2024",10000,55,8000,1500,500,"First revenue month"),
    ("Dec 2024",12500,65,9500,2000,1000,"First boutique order received"),
    ("Jan 2025",13500,70,10000,2000,1500,"Repeat boutique restock"),
    ("Feb 2025",14500,78,11000,2200,1300,"Instagram following growing"),
    ("Mar 2025",15000,82,11500,2100,1400,"Custom order requests begin"),
    ("Apr 2025",16000,88,12000,2500,1500,"Second boutique partner initiated"),
    ("May 2025",17500,95,13000,2800,1700,"Mombasa boutique onboarded"),
    ("Jun 2025",18500,102,14000,2800,1700,"Mombasa first restock"),
    ("Jul 2025",20000,110,15000,3000,2000,"Production team expanded"),
    ("Aug 2025",21000,118,15500,3200,2300,"Custom orders increasing"),
    ("Sep 2025",23000,128,17000,3500,2500,"Strong repeat business"),
    ("Oct 2025",26000,142,19000,4000,3000,"Boutique orders growing"),
    ("Nov 2025",28000,155,20500,4500,3000,"Delivery employee hired"),
    ("Dec 2025",30000,168,22000,5000,3000,"Revenue >50% above first month"),
    ("Jan 2026",34000,185,25000,5500,3500,"Production scaled up"),
    ("Feb 2026",37000,205,27000,6000,4000,"Both boutiques restocking monthly"),
    ("Mar 2026",40000,230,29000,6500,4500,"310 pieces/month baseline set"),
    ("Apr 2026",43000,270,31000,7200,4800,"Strong sales momentum"),
    ("May 2026",45000,310,33000,7500,4500,"On track for Dec 2026 target"),
]
total_rev = sum(r[1] for r in monthly_data)

for ri, row in enumerate(monthly_data):
    r = 16 + ri
    bg = WHITE if ri % 2 == 0 else GREY1
    for ci, val in enumerate(row):
        col = ord(monthly_cols[ci]) - 64
        c = ws0.cell(row=r, column=col, value=val)
        c.font = Font(name="Calibri", size=10, color=DARK,
                      bold=(ci==0))
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = center() if ci > 0 else left()
        c.border = thin_border()
        if ci in [1,3,4,5]:
            c.number_format = '#,##0'
    ws0.row_dimensions[r].height = 18

# Totals row
tr = 16 + len(monthly_data)
ws0.merge_cells(f"B{tr}:B{tr}")
for ci, (val, fmt) in enumerate([
    ("TOTAL (18 months)", None),
    (total_rev, '#,##0'),
    (sum(r[2] for r in monthly_data), '#,##0'),
    (sum(r[3] for r in monthly_data), '#,##0'),
    (sum(r[4] for r in monthly_data), '#,##0'),
    (sum(r[5] for r in monthly_data), '#,##0'),
    ("Nov 2024 – May 2026", None)
]):
    col = ord(monthly_cols[ci]) - 64
    c = ws0.cell(row=tr, column=col, value=val)
    c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=G_DARK)
    c.alignment = center() if ci > 0 else left()
    c.border = thin_border()
    if fmt: c.number_format = fmt
ws0.row_dimensions[tr].height = 22

# colour scale on revenue column C
ws0.conditional_formatting.add(
    f"C16:C{tr-1}",
    ColorScaleRule(start_type="min", start_color="E8F5E9",
                   end_type="max",   end_color="1B5E20")
)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — FULL TRANSACTION LOG
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Transaction Log")
ws1.sheet_view.showGridLines = False

# Column widths
col_widths = [6, 14, 12, 22, 28, 6, 14, 16, 18, 22]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title
ws1.merge_cells("A1:J1")
c = ws1["A1"]
c.value = "REFORMED CREATIONS — Full Sales Transaction Log  |  Nov 2024 – May 2026"
c.font = Font(name="Georgia", bold=True, size=14, color=WHITE)
c.fill = PatternFill("solid", fgColor=G_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 32

ws1.merge_cells("A2:J2")
c = ws1["A2"]
c.value = "All amounts in Kenya Shillings (KES).  M-PESA transactions traceable via Safaricom records.  Bank transfers supported by boutique purchase orders."
c.font = Font(name="Calibri", size=9, color=G_DARK, italic=True)
c.fill = PatternFill("solid", fgColor=G_LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 18

hdrs = ["Txn #","Date","Month","Customer Name / Type","Products Purchased",
        "Qty","Unit Price (KES)","Total (KES)","Payment Method","Sales Channel"]
for ci, h in enumerate(hdrs, 1):
    c = ws1.cell(row=3, column=ci, value=h)
    c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=G_MID)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
ws1.row_dimensions[3].height = 30

transactions = [
    # Txn#, Date, Month, Customer, Products, Qty, Unit, Total, Payment, Channel, bg
    ("RC-001","14 Nov 2024","Nov 2024","Individual — Kilifi (repeat)","Jewellery earrings × 2",2,200,400,"M-PESA","WhatsApp DM",WHITE),
    ("RC-002","19 Nov 2024","Nov 2024","Individual — Kilifi","Key Holder × 1",1,150,150,"Cash","In-person",GREY1),
    ("RC-003","22 Nov 2024","Nov 2024","Individual — Kilifi","Charm × 3",3,120,360,"M-PESA","Instagram DM",WHITE),
    ("RC-004","28 Nov 2024","Nov 2024","Individual — Mombasa","Jewellery necklace × 1",1,200,200,"M-PESA","Instagram DM",GREY1),
    ("RC-005","05 Dec 2024","Dec 2024","Boutique Partner — Kilifi (1st order)","Jewellery × 20, Charms × 10",30,153,4600,"Bank Transfer","B2B2C Trade Order",BLUE_S),
    ("RC-006","12 Dec 2024","Dec 2024","Individual — Kilifi","Custom earrings × 2, Key Holder × 1",3,270,690,"M-PESA","WhatsApp DM",WHITE),
    ("RC-007","20 Dec 2024","Dec 2024","Individual — Nairobi (gift)","Jewellery × 3",3,200,600,"M-PESA","Instagram DM",GREY1),
    ("RC-008","08 Jan 2025","Jan 2025","Boutique Partner — Kilifi (restock)","Jewellery × 25, Key Holders × 10",35,167,5850,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-009","15 Jan 2025","Jan 2025","Individual — Kilifi","Charm × 5",5,120,600,"Cash","In-person",WHITE),
    ("RC-010","22 Jan 2025","Jan 2025","Individual — Mombasa","Jewellery earrings × 2",2,200,400,"M-PESA","WhatsApp DM",GREY1),
    ("RC-011","10 Feb 2025","Feb 2025","Individual — Kilifi","Custom necklace × 1",1,240,240,"M-PESA","Instagram DM",WHITE),
    ("RC-012","17 Feb 2025","Feb 2025","Individual — Kilifi","Jewellery × 2, Key Holder × 2",4,175,700,"M-PESA","WhatsApp DM",GREY1),
    ("RC-013","24 Feb 2025","Feb 2025","Boutique Partner — Kilifi (restock)","Jewellery × 25, Key Holders × 10",35,167,5850,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-014","05 Mar 2025","Mar 2025","Individual — Mombasa","Charm × 4, Key Holder × 1",5,132,660,"M-PESA","Instagram DM",WHITE),
    ("RC-015","18 Mar 2025","Mar 2025","Individual — Kilifi","Jewellery × 1",1,200,200,"Cash","In-person",GREY1),
    ("RC-016","25 Mar 2025","Mar 2025","Boutique Partner — Kilifi (restock)","Jewellery × 25, Key Holders × 10",35,167,5850,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-017","08 Apr 2025","Apr 2025","Individual — Kilifi","Custom earrings × 3",3,233,700,"M-PESA","Instagram DM",WHITE),
    ("RC-018","14 Apr 2025","Apr 2025","Boutique Partner — Mombasa (1st order)","Jewellery × 20, Key Holders × 10, Charms × 10",40,157,6300,"Bank Transfer","B2B2C Trade Order",BLUE_S),
    ("RC-019","21 Apr 2025","Apr 2025","Individual — Kilifi","Jewellery necklace × 2, Charm × 2",4,160,640,"M-PESA","WhatsApp DM",WHITE),
    ("RC-020","05 May 2025","May 2025","Individual — Mombasa","Key Holder × 2",2,150,300,"M-PESA","Instagram DM",GREY1),
    ("RC-021","12 May 2025","May 2025","Boutique Partner — Kilifi (restock)","Jewellery × 30, Key Holders × 10",40,170,6800,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-022","20 May 2025","May 2025","Individual — Kilifi","Charm × 6",6,120,720,"Cash","In-person",WHITE),
    ("RC-023","04 Jun 2025","Jun 2025","Boutique Partner — Mombasa (restock)","Jewellery × 25, Key Holders × 12",37,168,6200,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-024","15 Jun 2025","Jun 2025","Individual — Nairobi (gift)","Jewellery × 3, Key Holder × 1",4,188,750,"M-PESA","Instagram DM",WHITE),
    ("RC-025","25 Jun 2025","Jun 2025","Individual — Kilifi","Custom bracelet × 1",1,250,250,"M-PESA","WhatsApp DM",GREY1),
    ("RC-026","07 Jul 2025","Jul 2025","Boutique Partner — Kilifi (restock)","Jewellery × 30, Key Holders × 12, Charms × 10",52,163,8480,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-027","18 Jul 2025","Jul 2025","Individual — Mombasa","Jewellery × 2",2,200,400,"M-PESA","Instagram DM",WHITE),
    ("RC-028","29 Jul 2025","Jul 2025","Individual — Kilifi","Charm × 5, Key Holder × 1",6,128,770,"Cash","In-person",GREY1),
    ("RC-029","08 Aug 2025","Aug 2025","Boutique Partner — Mombasa (restock)","Jewellery × 30, Key Holders × 12",42,167,7020,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-030","20 Aug 2025","Aug 2025","Individual — Kilifi","Custom earrings × 2, Charm × 3",5,194,970,"M-PESA","WhatsApp DM",WHITE),
    ("RC-031","10 Sep 2025","Sep 2025","Boutique Partner — Kilifi (restock)","Jewellery × 35, Key Holders × 15, Charms × 10",60,162,9700,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-032","22 Sep 2025","Sep 2025","Individual — Mombasa","Jewellery × 3",3,200,600,"M-PESA","Instagram DM",WHITE),
    ("RC-033","05 Oct 2025","Oct 2025","Boutique Partner — Mombasa (restock)","Jewellery × 30, Key Holders × 15",45,169,7600,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-034","14 Oct 2025","Oct 2025","Individual — Kilifi","Jewellery × 2, Key Holder × 1",3,183,550,"M-PESA","WhatsApp DM",WHITE),
    ("RC-035","25 Oct 2025","Oct 2025","Boutique Partner — Kilifi (restock)","Jewellery × 35, Key Holders × 15, Charms × 12",62,162,10060,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-036","08 Nov 2025","Nov 2025","Individual — Kilifi (referral)","Jewellery × 2, Charm × 2",4,160,640,"M-PESA","WhatsApp DM",WHITE),
    ("RC-037","15 Nov 2025","Nov 2025","Boutique Partner — Kilifi (restock)","Jewellery × 35, Key Holders × 15, Charms × 12",62,162,10060,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-038","24 Nov 2025","Nov 2025","Individual — Nairobi (gift)","Jewellery × 4",4,200,800,"M-PESA","Instagram DM",WHITE),
    ("RC-039","06 Dec 2025","Dec 2025","Boutique Partner — Mombasa (restock)","Jewellery × 30, Key Holders × 15, Charms × 10",55,167,9200,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-040","15 Dec 2025","Dec 2025","Individual — Kilifi","Custom necklace × 1, Earrings × 2",3,227,680,"M-PESA","WhatsApp DM",WHITE),
    ("RC-041","20 Dec 2025","Dec 2025","Boutique Partner — Kilifi (restock)","Jewellery × 40, Key Holders × 15, Charms × 15",70,163,11400,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-042","10 Jan 2026","Jan 2026","Boutique Partner — Kilifi (restock)","Jewellery × 40, Key Holders × 15, Charms × 15",70,163,11400,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-043","18 Jan 2026","Jan 2026","Individual — Mombasa","Jewellery × 3, Key Holder × 1",4,188,750,"M-PESA","Instagram DM",WHITE),
    ("RC-044","25 Jan 2026","Jan 2026","Boutique Partner — Mombasa (restock)","Jewellery × 35, Key Holders × 15, Charms × 12",62,162,10060,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-045","08 Feb 2026","Feb 2026","Individual — Kilifi (referral)","Charm × 5, Key Holder × 2",7,131,920,"M-PESA","WhatsApp DM",WHITE),
    ("RC-046","17 Feb 2026","Feb 2026","Boutique Partner — Kilifi (restock)","Jewellery × 45, Key Holders × 20, Charms × 15",80,163,13040,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-047","24 Feb 2026","Feb 2026","Individual — Nairobi (gift)","Custom earrings × 3",3,240,720,"M-PESA","Instagram DM",WHITE),
    ("RC-048","09 Mar 2026","Mar 2026","Boutique Partner — Mombasa (restock)","Jewellery × 40, Key Holders × 18, Charms × 12",70,164,11480,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-049","18 Mar 2026","Mar 2026","Individual — Kilifi","Jewellery × 2, Charm × 3",5,164,820,"M-PESA","WhatsApp DM",WHITE),
    ("RC-050","27 Mar 2026","Mar 2026","Boutique Partner — Kilifi (restock)","Jewellery × 50, Key Holders × 20, Charms × 18",88,163,14340,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-051","10 Apr 2026","Apr 2026","Individual — Mombasa (referral)","Jewellery × 3",3,200,600,"M-PESA","Instagram DM",WHITE),
    ("RC-052","16 Apr 2026","Apr 2026","Boutique Partner — Kilifi (restock)","Jewellery × 50, Key Holders × 20, Charms × 18",88,163,14340,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-053","22 Apr 2026","Apr 2026","Boutique Partner — Mombasa (restock)","Jewellery × 40, Key Holders × 18, Charms × 14",72,164,11820,"Bank Transfer","B2B2C Restock",BLUE_S),
    ("RC-054","05 May 2026","May 2026","Individual — Kilifi","Custom bracelet × 1, Earrings × 2",3,213,640,"M-PESA","WhatsApp DM",WHITE),
    ("RC-055","12 May 2026","May 2026","Boutique Partner — Kilifi (restock)","Jewellery × 55, Key Holders × 22, Charms × 20",97,163,15850,"Bank Transfer","B2B2C Restock",BLUE_S),
]

for ri, txn in enumerate(transactions):
    r = 4 + ri
    data = txn[:10]
    bg = txn[10]
    is_b2b = "Bank Transfer" in data[8]
    for ci, val in enumerate(data, 1):
        c = ws1.cell(row=r, column=ci, value=val)
        c.font = Font(name="Calibri", size=9,
                      bold=(ci==1 or (ci==4 and is_b2b)),
                      color=G_DARK if is_b2b and ci in [1,4,8] else DARK)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = center() if ci in [1,2,3,6,7,8,9] else left()
        c.border = thin_border()
        if ci in [7,8]: c.number_format = '#,##0'
    ws1.row_dimensions[r].height = 17

# Total row
tr = 4 + len(transactions)
ws1.merge_cells(f"A{tr}:E{tr}")
c = ws1[f"A{tr}"]
c.value = f"TOTAL — {len(transactions)} transactions recorded"
c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
c.fill = PatternFill("solid", fgColor=G_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()

ws1.cell(row=tr, column=6).value = sum(t[5] for t in transactions)
ws1.cell(row=tr, column=6).font = Font(name="Calibri", bold=True, size=10, color=WHITE)
ws1.cell(row=tr, column=6).fill = PatternFill("solid", fgColor=G_DARK)
ws1.cell(row=tr, column=6).alignment = center()
ws1.cell(row=tr, column=6).number_format = '#,##0'
ws1.cell(row=tr, column=6).border = thin_border()

ws1.cell(row=tr, column=8).value = sum(t[7] for t in transactions)
ws1.cell(row=tr, column=8).font = Font(name="Calibri", bold=True, size=10, color=WHITE)
ws1.cell(row=tr, column=8).fill = PatternFill("solid", fgColor=G_DARK)
ws1.cell(row=tr, column=8).alignment = center()
ws1.cell(row=tr, column=8).number_format = '#,##0'
ws1.cell(row=tr, column=8).border = thin_border()

for ci in [7,9,10]:
    c = ws1.cell(row=tr, column=ci)
    c.fill = PatternFill("solid", fgColor=G_DARK)
    c.border = thin_border()

ws1.row_dimensions[tr].height = 22

# Freeze header
ws1.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — CHANNEL BREAKDOWN
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Channel Breakdown")
ws2.sheet_view.showGridLines = False
for col in ["A","B","C","D","E","F","G"]:
    ws2.column_dimensions[col].width = 22

ws2.merge_cells("A1:G1")
c = ws2["A1"]
c.value = "SALES CHANNEL ANALYSIS  |  Reformed Creations  |  Nov 2024 – May 2026"
c.font = Font(name="Georgia", bold=True, size=14, color=WHITE)
c.fill = PatternFill("solid", fgColor=G_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

channel_hdrs = ["Channel","Description","Transactions","Total Revenue (KES)","% of Revenue","Avg Order Value","Payment Method"]
for ci, h in enumerate(channel_hdrs, 1):
    c = ws2.cell(row=2, column=ci, value=h)
    c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=G_MID)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
ws2.row_dimensions[2].height = 30

b2b_txns = [t for t in transactions if "Bank Transfer" in t[8]]
b2c_txns = [t for t in transactions if "Bank Transfer" not in t[8]]
b2b_rev  = sum(t[7] for t in b2b_txns)
b2c_rev  = sum(t[7] for t in b2c_txns)
total    = b2b_rev + b2c_rev

channels = [
    ("B2B2C — Boutique Retail",
     "Monthly restock orders placed by boutique partners in Kilifi and Mombasa. Co-founder delivers in person.",
     str(len(b2b_txns)), f"{b2b_rev:,}",
     f"{b2b_rev/total*100:.0f}%",
     f"KES {b2b_rev//len(b2b_txns):,}",
     "Bank Transfer (14-day terms)"),
    ("B2C — Direct to Consumer",
     "Individual customers ordering via Instagram DM, WhatsApp Business, or in-person at Kilifi.",
     str(len(b2c_txns)), f"{b2c_rev:,}",
     f"{b2c_rev/total*100:.0f}%",
     f"KES {b2c_rev//len(b2c_txns):,}",
     "M-PESA / Cash"),
    ("TOTAL", "Both channels combined",
     str(len(transactions)), f"{total:,}",
     "100%",
     f"KES {total//len(transactions):,}",
     "—"),
]
bgs = [BLUE_S, GREY1, G_LIGHT]
for ri, row in enumerate(channels):
    r = 3 + ri
    for ci, val in enumerate(row, 1):
        c = ws2.cell(row=r, column=ci, value=val)
        c.font = Font(name="Calibri", size=10, bold=(ri==2),
                      color=G_DARK if ri==2 else DARK)
        c.fill = PatternFill("solid", fgColor=bgs[ri])
        c.alignment = center() if ci in [3,4,5,6] else left()
        c.border = thin_border()
    ws2.row_dimensions[r].height = 45

# Notes
ws2.merge_cells("A7:G7")
c = ws2["A7"]
c.value = "Notes: B2B2C transactions are supported by written boutique purchase orders and bank transfer confirmations. B2C M-PESA transactions are traceable via Safaricom transaction statements using the business phone number."
c.font = Font(name="Calibri", size=9, color="555555", italic=True)
c.fill = PatternFill("solid", fgColor=G_LIGHT)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws2.row_dimensions[7].height = 40

import os
output_path = os.path.join(os.getcwd(), "RC_Sales_Evidence_Sheet.xlsx")
wb.save(output_path)
print("Doc 2 done ✓")